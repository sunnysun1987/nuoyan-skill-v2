import hashlib
import json
from pathlib import Path

import pytest
from bs4 import BeautifulSoup

from ivd_research.evidence import (
    build_draft_evidence_card,
    export_evidence_card_files,
    generate_draft_evidence_cards,
)
from ivd_research.jsonl import append_jsonl, read_jsonl, write_json
from ivd_research.models import Material
from ivd_research.package import (
    build_standard_delivery,
    requires_life_science_research,
    scenario_coverage_warnings,
    verify_package,
)
from ivd_research.project_profile import formal_scenarios_for
from ivd_research.review_excel import export_review, import_review
from ivd_research.research_integrity import (
    record_research_claim,
    record_research_iteration,
)
from ivd_research.source_adapters.life_science_research_bridge import import_life_science_findings
from ivd_research.status import init_task
from ivd_research.confirmations import update_confirmations


FULL_CONFIRMATIONS = {
    "task_info": True,
    "keyword_pool": True,
    "collection_scope": True,
    "primary_query": "血浆 p-tau217 阿尔茨海默病 体外诊断",
    "english_keywords": "plasma p-tau217 Alzheimer disease IVD",
    "sample_type": "血浆",
    "platform": "化学发光",
    "methodology": True,
    "intended_use": "阿尔茨海默病辅助诊断",
    "target_region": "中国",
    "competitor_scope": "NMPA 已注册同类产品",
    "patent_scope": True,
}

HCG_CONFIRMATIONS = {
    "task_info": True,
    "keyword_pool": True,
    "collection_scope": True,
    "primary_query": "beta-hCG定量检测试剂盒（荧光免疫层析法）",
    "english_keywords": "beta hCG quantitative test kit fluorescence immunochromatography",
    "sample_type": "血清/尿液",
    "platform": "荧光免疫层析",
    "methodology": "荧光免疫层析法",
    "intended_use": "妊娠相关检测",
    "target_region": "中国",
    "competitor_scope": "NMPA 已注册 hCG 同类产品",
    "patent_scope": "中国",
}


def _task_dir(tmp_path: Path) -> Path:
    state = init_task("p-tau217 稳定性测试", tmp_path)
    return Path(state.task_dir)


def test_build_standard_delivery_does_not_run_translation(monkeypatch, tmp_path: Path):
    task_dir = _task_dir(tmp_path)

    def forbidden_translation(*args, **kwargs):
        raise AssertionError("delivery rendering must only read translation cache")

    monkeypatch.setattr("ivd_research.translation.translate_materials", forbidden_translation)

    result = build_standard_delivery(task_dir)

    assert Path(result["delivery_dir"]).exists()


def _write_single_material_and_card(task_dir: Path) -> None:
    text_path = task_dir / "extracted_text" / "literature" / "MAT-000001.txt"
    text_path.parent.mkdir(parents=True, exist_ok=True)
    text_path.write_text(
        "PMID：12345678\nDOI：10.1000/test\n摘要：p-tau217 与 AD 病理相关。",
        encoding="utf-8",
    )
    material = Material(
        material_id="MAT-000001",
        task_id="TEST",
        source_scenario="pubmed_literature",
        material_type="literature",
        title="Plasma p-tau217 for Alzheimer disease diagnosis",
        source_url="https://pubmed.ncbi.nlm.nih.gov/12345678/",
        search_keyword_or_query="plasma p-tau217 Alzheimer disease",
        collection_path={"scenario_id": "pubmed_literature"},
        collection_time="2026-06-16T00:00:00+08:00",
        adapter_id="pubmed_literature",
        adapter_version="2.0.0",
        raw_fields={
            "pmid": "12345678",
            "doi": "10.1000/test",
            "journal": "Journal of Test Medicine",
            "publication_date": "2026-06-16",
            "abstract": "p-tau217 is associated with Alzheimer pathology.",
            "fulltext_status": "completed",
            "pdf_status": "not_available",
        },
        extracted_text_status="completed",
        extracted_text_path=str(text_path.relative_to(task_dir)),
    )
    append_jsonl(task_dir / "data" / "materials.jsonl", material.model_dump(mode="json"))
    card = build_draft_evidence_card(
        task_dir,
        material.model_dump(mode="json"),
        "EC-000001",
    )
    card_payload = card.model_dump(mode="json")
    append_jsonl(task_dir / "data" / "evidence_cards.jsonl", card_payload)
    export_evidence_card_files(task_dir, card_payload)


def _write_hcg_material_and_card(task_dir: Path) -> None:
    text_path = task_dir / "extracted_text" / "literature" / "MAT-000001.txt"
    text_path.parent.mkdir(parents=True, exist_ok=True)
    text_path.write_text(
        "摘要：human chorionic gonadotropin assays support pregnancy-related testing.",
        encoding="utf-8",
    )
    material = Material(
        material_id="MAT-000001",
        task_id="TEST",
        source_scenario="pubmed_literature",
        material_type="literature",
        title="Human chorionic gonadotropin immunoassay for pregnancy-related testing",
        source_url="https://pubmed.ncbi.nlm.nih.gov/12345678/",
        search_keyword_or_query="beta hCG quantitative immunoassay",
        collection_path={"scenario_id": "pubmed_literature"},
        collection_time="2026-07-09T00:00:00+08:00",
        adapter_id="pubmed_literature",
        adapter_version="2.0.0",
        raw_fields={
            "pmid": "12345678",
            "journal": "Journal of Test Medicine",
            "publication_date": "2026-07-09",
            "abstract": "Human chorionic gonadotropin assays support pregnancy-related testing.",
            "fulltext_status": "completed",
        },
        extracted_text_status="completed",
        extracted_text_path=str(text_path.relative_to(task_dir)),
    )
    append_jsonl(task_dir / "data" / "materials.jsonl", material.model_dump(mode="json"))
    card = build_draft_evidence_card(
        task_dir,
        material.model_dump(mode="json"),
        "EC-000001",
    )
    card_payload = card.model_dump(mode="json")
    append_jsonl(task_dir / "data" / "evidence_cards.jsonl", card_payload)
    export_evidence_card_files(task_dir, card_payload)


def _complete_research_integrity(task_dir: Path) -> None:
    reviewed_cards = list(read_jsonl(task_dir / "data" / "reviewed_evidence_cards.jsonl"))
    included_ids = [
        card["evidence_card_id"]
        for card in reviewed_cards
        if card.get("include_in_report")
    ]
    record_research_claim(
        task_dir,
        {
            "claim_id": "CLM-000001",
            "text": "现有证据支持继续开展研发验证。",
            "claim_type": "research_judgement",
            "status": "supported",
            "evidence_card_ids": included_ids,
            "confidence": "medium",
            "needs_human_review": False,
            "reviewed_at": "2026-08-04T00:00:00+08:00",
            "reviewer_role": "研发专家",
        },
    )
    for iteration_id, direction, axis in (
        ("ITER-001", "反向证据审计", "claim_stance"),
        ("ITER-002", "覆盖缺口审计", "coverage_gap"),
    ):
        record_research_iteration(
            task_dir,
            {
                "iteration_id": iteration_id,
                "direction": direction,
                "pivot_axis": axis,
                "quality_targets_met": True,
            },
        )


def _mark_nmpa_verified_zero(task_dir: Path, task: dict) -> None:
    attempt_id = "NMPA-TEST-001"
    plan_path = task_dir / "manual" / "nmpa" / "search_plan.json"
    record_path = task_dir / "manual" / "nmpa" / "search_records" / "TEST.json"
    manifest_path = task_dir / "manual" / "nmpa" / "import_manifests" / "TEST.json"
    evidence_path = (
        task_dir
        / "downloads"
        / "competitors"
        / "nmpa_manual"
        / "TEST"
        / attempt_id
        / "001_zero-result.png"
    )
    evidence_path.parent.mkdir(parents=True, exist_ok=True)
    evidence_path.write_bytes(b"\x89PNG\r\n\x1a\nverified zero result")
    evidence_sha256 = hashlib.sha256(evidence_path.read_bytes()).hexdigest()
    official_url = (
        "https://www.nmpa.gov.cn/datasearch/home-index.html#category=ylqx"
    )
    write_json(
        plan_path,
        {
            "task_id": task["task_id"],
            "attempts": [
                {
                    "attempt_id": attempt_id,
                    "query": "p-tau217",
                    "registration_type": "境内医疗器械（注册）",
                    "official_search_url": official_url,
                }
            ],
        },
    )
    write_json(
        record_path,
        {
            "task_id": task["task_id"],
            "search_session_id": "TEST",
            "operator_confirmed": True,
            "attempts": [
                {
                    "attempt_id": attempt_id,
                    "query": "p-tau217",
                    "registration_type": "境内医疗器械（注册）",
                    "official_search_url": official_url,
                    "search_time": "2026-07-23T10:00:00+08:00",
                    "result_count": 0,
                }
            ],
        },
    )
    write_json(
        manifest_path,
        {
            "task_id": task["task_id"],
            "search_session_id": "TEST",
            "capture_complete": True,
            "attempts": [
                {
                    "attempt_id": attempt_id,
                    "evidence_files": [
                        {
                            "relative_path": evidence_path.relative_to(task_dir).as_posix(),
                            "sha256": evidence_sha256,
                            "status": "imported",
                        }
                    ],
                    "results": [],
                }
            ],
        },
    )
    scenario = task["scenario_statuses"]["nmpa_competitor"]
    scenario["status"] = "no_results"
    scenario["last_message"] = "离线测试：NMPA 零结果已完成人工证据核验。"
    scenario["manual_collection"] = {
        "phase": "verified_no_results",
        "plan_path": plan_path.relative_to(task_dir).as_posix(),
        "plan_sha256": hashlib.sha256(plan_path.read_bytes()).hexdigest(),
        "search_record_path": record_path.relative_to(task_dir).as_posix(),
        "search_record_sha256": hashlib.sha256(record_path.read_bytes()).hexdigest(),
        "manifest_path": manifest_path.relative_to(task_dir).as_posix(),
        "manifest_sha256": hashlib.sha256(manifest_path.read_bytes()).hexdigest(),
        "required_attempt_ids": [attempt_id],
        "recorded_attempt_ids": [attempt_id],
        "validated_attempt_ids": [attempt_id],
        "imported_material_ids": [],
        "observed_result_count": 0,
        "zero_results_verified": True,
        "last_updated": "2026-07-23T10:01:00+08:00",
    }


def _mark_formal_scenarios(task_dir: Path, status: str = "no_results") -> None:
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    for scenario_id in formal_scenarios_for(task):
        task["scenario_statuses"][scenario_id]["status"] = status
        task["scenario_statuses"][scenario_id]["last_message"] = "离线测试：已记录明确状态。"
    task["scenario_statuses"]["pubmed_literature"]["status"] = "completed"
    task["scenario_statuses"]["pubmed_literature"]["material_count"] = 1
    if status == "no_results":
        _mark_nmpa_verified_zero(task_dir, task)
    write_json(task_dir / "task.json", task)
    query_attempts = {
        "cmde_regulatory": [
            ("core_cn", "p-tau217"),
            ("broad_cn", "血浆 p-tau217 阿尔茨海默病"),
        ],
        "standards_current": [
            ("core_cn", "p-tau217"),
            ("core_product_cn", "p-tau217 测定试剂盒"),
        ],
        "openalex_literature": [
            ("openalex_core_keywords", "plasma p-tau217 Alzheimer disease immunoassay"),
            ("openalex_broad_keywords", "plasma p-tau217 Alzheimer disease IVD"),
        ],
        "yiigle_fulltext": [
            ("yiigle_fulltext_core_expression", "篇关摘=(p-tau217)"),
            ("yiigle_fulltext_expression", "篇关摘=(血浆 p-tau217 阿尔茨海默病)"),
        ],
        "yiigle_zhjyyxzz": [
            ("core_cn", "p-tau217"),
            ("broad_cn", "血浆 p-tau217 阿尔茨海默病"),
        ],
        "cma_lab_management": [
            ("core_cn", "p-tau217"),
            ("broad_cn", "血浆 p-tau217 阿尔茨海默病"),
        ],
        "yiigle_zhsjkzz": [
            ("core_cn", "p-tau217"),
            ("broad_cn", "血浆 p-tau217 阿尔茨海默病"),
        ],
    }
    for scenario_id, attempts in query_attempts.items():
        if scenario_id not in {item for item in formal_scenarios_for(task)}:
            continue
        append_jsonl(
            task_dir / "logs" / "events.jsonl",
            {
                "event": "scenario_query_attempts",
                "scenario_id": scenario_id,
                "attempts": [
                    {
                        "query_role": role,
                        "query": query,
                        "status": status,
                        "material_count": 0,
                        "message_zh": "离线测试：核心检索层级已记录。",
                    }
                    for role, query in attempts
                ],
            },
        )


def _import_complete_life_science_coverage(task_dir: Path) -> None:
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    findings = []
    seed_rows = [
        ("UniProt", "protein", "MAPT"),
        ("Open Targets", "target", "MAPT"),
        ("ClinicalTrials.gov", "clinical", "p-tau217"),
        ("ClinicalTrials.gov", "clinical", "p-tau181"),
        ("STRING", "network", "MAPT"),
        ("Reactome", "pathway", "tau protein binding"),
        ("QuickGO", "pathway", "microtubule binding"),
        ("Human Protein Atlas", "protein", "MAPT brain expression"),
        ("GWAS Catalog", "genetics", "Alzheimer disease"),
        ("NCBI Gene", "target", "MAPT"),
        ("EFO/OLS", "disease", "Alzheimer disease"),
        ("ClinicalTrials.gov", "clinical", "plasma biomarker"),
    ]
    for index, (database, lane, entity) in enumerate(seed_rows, start=1):
        findings.append(
            {
                "source_database": database,
                "evidence_lane": lane,
                "entity": entity,
                "query": "plasma p-tau217 Alzheimer disease",
                "result_summary": f"{database} evidence row {index} for {entity}.",
                "source_url": f"https://example.org/life-science/{index}",
                "identifier": f"LS-{index:03d}",
            }
        )
    import_life_science_findings(
        task["task_id"],
        task_dir,
        findings,
        query="plasma p-tau217 Alzheimer disease",
    )


def test_verify_package_keeps_incomplete_scope_as_not_business_ready(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)

    result = verify_package(task_dir)

    assert result["delivery_artifacts_ready"] is True
    assert (task_dir / "交付目录" / "02_证据卡" / "EC-000001.md").exists()
    assert result["business_ready"] is False
    assert "task_info" in result["missing_confirmations"]
    assert result["final_review_ready"] is False


def test_unverified_nmpa_no_results_blocks_scenario_coverage(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    scenario = task["scenario_statuses"]["nmpa_competitor"]
    scenario["status"] = "no_results"
    scenario["last_message"] = "人工声称无结果，但未提交证据。"
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "未经人工证据核验" in warning for warning in warnings)


def test_nmpa_completed_without_dedicated_manifest_blocks_coverage(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    scenario = task["scenario_statuses"]["nmpa_competitor"]
    scenario["status"] = "completed"
    scenario["material_count"] = 1
    scenario["last_message"] = "通过通用 import-finding 导入一条线索。"
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "专用人工导入" in warning for warning in warnings)


def test_verified_nmpa_zero_results_satisfies_source_coverage(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    _mark_nmpa_verified_zero(task_dir, task)
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert not any("NMPA" in warning for warning in warnings)


def test_tampered_nmpa_evidence_blocks_verified_zero_results(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    _mark_nmpa_verified_zero(task_dir, task)
    write_json(task_dir / "task.json", task)
    manual = task["scenario_statuses"]["nmpa_competitor"]["manual_collection"]
    manifest = json.loads((task_dir / manual["manifest_path"]).read_text(encoding="utf-8"))
    evidence_path = task_dir / manifest["attempts"][0]["evidence_files"][0][
        "relative_path"
    ]
    evidence_path.write_bytes(b"tampered")

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "校验值" in warning for warning in warnings)


def test_rehashed_nmpa_record_mismatch_still_blocks_verified_zero_results(
    tmp_path: Path,
):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    _mark_nmpa_verified_zero(task_dir, task)
    manual = task["scenario_statuses"]["nmpa_competitor"]["manual_collection"]
    record_path = task_dir / manual["search_record_path"]
    record = json.loads(record_path.read_text(encoding="utf-8"))
    record["attempts"][0]["query"] = "与检索计划不一致的查询"
    write_json(record_path, record)
    manual["search_record_sha256"] = hashlib.sha256(record_path.read_bytes()).hexdigest()
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "检索计划不一致" in warning for warning in warnings)


def test_rehashed_invalid_nmpa_evidence_format_still_blocks_verified_zero_results(
    tmp_path: Path,
):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    _mark_nmpa_verified_zero(task_dir, task)
    manual = task["scenario_statuses"]["nmpa_competitor"]["manual_collection"]
    manifest_path = task_dir / manual["manifest_path"]
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    evidence = manifest["attempts"][0]["evidence_files"][0]
    evidence_path = task_dir / evidence["relative_path"]
    evidence_path.write_bytes(b"not a png")
    evidence["sha256"] = hashlib.sha256(evidence_path.read_bytes()).hexdigest()
    write_json(manifest_path, manifest)
    manual["manifest_sha256"] = hashlib.sha256(manifest_path.read_bytes()).hexdigest()
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "格式" in warning for warning in warnings)


def test_malformed_nmpa_artifact_fails_closed_without_crashing(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    _mark_nmpa_verified_zero(task_dir, task)
    manual = task["scenario_statuses"]["nmpa_competitor"]["manual_collection"]
    plan_path = task_dir / manual["plan_path"]
    plan = json.loads(plan_path.read_text(encoding="utf-8"))
    plan["attempts"] = ["broken-attempt"]
    write_json(plan_path, plan)
    manual["plan_sha256"] = hashlib.sha256(plan_path.read_bytes()).hexdigest()
    manual["observed_result_count"] = "not-an-integer"
    write_json(task_dir / "task.json", task)

    warnings = scenario_coverage_warnings(task_dir)

    assert any("NMPA" in warning and "无效" in warning for warning in warnings)


def test_life_science_requirement_detects_ad_without_lead_false_positive():
    assert requires_life_science_research(
        {
            "topic": "AD p-tau181 血液标志物",
            "confirmations": {"english_keywords": "AD plasma biomarker"},
        }
    )
    assert not requires_life_science_research(
        {
            "topic": "lead time workflow improvement",
            "confirmations": {"english_keywords": "lead time process validation"},
        }
    )


def test_standard_delivery_report_has_drilldown_navigation_and_metric_definitions(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)

    html = (task_dir / "交付目录" / "00_立项调研综合报告.html").read_text(encoding="utf-8")

    assert "章节目录" in html
    assert "analysis-nav-card" in html
    assert "点击后跳转到对应章节" in html
    assert 'data-jump-tab="tab-screening"' in html
    assert 'data-jump-target="screening-card-list"' in html
    assert 'data-tab-target="tab-metrics">指标事实' in html
    assert 'data-jump-tab="tab-metrics" data-jump-target="metric-facts"' in html
    assert 'data-jump-target="metric-facts"' in html
    assert 'id="tab-metrics"' in html
    assert "已写入材料库的原始资料总数" in html
    assert "从证据中结构化抽取的样本量" in html
    assert "口径：" not in html
    assert "中文阅读覆盖" not in html
    assert "metric-global-search" in html
    assert 'data-metric-filter="metric"' in html
    assert "指标字段" in html
    assert "数值字段" in html
    assert "analysis-evidence-table" in html
    assert "analysis-evidence-panel" in html
    assert 'data-page-size="8"' in html
    assert "依据清单" in html
    assert "相关内容原文" in html
    assert "先看结论" in html
    assert "研发定位" in html
    assert html.index('id="tab-metrics"') < html.index('id="tab-core"')

    soup = BeautifulSoup(html, "html.parser")
    reading_panel = soup.select_one("#tab-reading")
    metrics_panel = soup.select_one("#tab-metrics")
    assert reading_panel is not None
    assert metrics_panel is not None
    assert reading_panel.select_one("#metric-facts") is None
    assert metrics_panel.select_one("#metric-facts") is not None


def test_standard_delivery_merges_validated_report_sections_without_losing_workbench_tabs(
    tmp_path: Path,
):
    task_dir = _task_dir(tmp_path)
    _write_single_material_and_card(task_dir)
    append_jsonl(
        task_dir / "data" / "report_sections.jsonl",
        {
            "section_id": "SEC-01",
            "section_title": "临床意义",
            "facts": ["RSV 项目事实证据应出现在标准综合报告中。"],
            "analysis": "这是基于项目材料生成的专项分析，不是规则模板结论。",
            "evidence_gaps": ["仍需临床负责人复核适用人群。"],
            "evidence_strength_summary": "strong",
            "confidence_level": "高",
            "supporting_evidence_refs": [
                {
                    "material_id": "MAT-000001",
                    "evidence_card_id": "EC-000001",
                    "excerpt": "p-tau217 is associated with Alzheimer pathology.",
                }
            ],
            "needs_human_review": True,
        },
    )
    export_review(task_dir)
    build_standard_delivery(task_dir)

    html = (task_dir / "交付目录" / "00_立项调研综合报告.html").read_text(
        encoding="utf-8"
    )
    soup = BeautifulSoup(html, "html.parser")
    analysis_card = soup.select_one("#analysis-1")

    assert analysis_card is not None
    assert "这是基于项目材料生成的专项分析" in analysis_card.get_text(" ", strip=True)
    assert "RSV 项目事实证据应出现在标准综合报告中" in analysis_card.get_text(" ", strip=True)
    assert "证据强度：强" in analysis_card.get_text(" ", strip=True)
    assert "可信度：高" in analysis_card.get_text(" ", strip=True)
    assert "待人工复核" in analysis_card.get_text(" ", strip=True)
    assert len(analysis_card.select("tr[data-page-row]")) == 1
    assert 'data-jump-target="evidence-card-EC-000001"' in str(analysis_card)
    for panel_id in (
        "tab-analysis",
        "tab-reading",
        "tab-metrics",
        "tab-core",
        "tab-screening",
        "tab-gaps",
    ):
        assert soup.select_one(f"#{panel_id}") is not None


def test_all_evidence_tab_keeps_relevance_exclusions_visible_for_audit(
    tmp_path: Path,
):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    excluded = Material(
        material_id="MAT-000002",
        task_id="TEST",
        source_scenario="standards_current",
        material_type="standard",
        title="禽腺病毒4型荧光定量PCR检测方法",
        source_url="https://example.org/animal-standard",
        collection_time="2026-08-05T00:00:00+08:00",
        raw_fields={"standard_name": "禽腺病毒4型荧光定量PCR检测方法", "trade": "畜牧业"},
    )
    append_jsonl(
        task_dir / "data" / "materials.jsonl",
        excluded.model_dump(mode="json"),
    )
    excluded_card = build_draft_evidence_card(
        task_dir,
        excluded.model_dump(mode="json"),
        "EC-000002",
    )
    append_jsonl(
        task_dir / "data" / "evidence_cards.jsonl",
        excluded_card.model_dump(mode="json"),
    )

    build_standard_delivery(task_dir)

    html = (task_dir / "交付目录" / "00_立项调研综合报告.html").read_text(
        encoding="utf-8"
    )
    soup = BeautifulSoup(html, "html.parser")
    panel = soup.select_one("#tab-screening")

    assert panel is not None
    assert len(panel.select(".filter-card")) == 2
    assert panel.select_one('[data-relevance="相关性排除"]') is not None
    assert "相关性排除" in panel.get_text(" ", strip=True)
    assert "2 / 2" in panel.get_text(" ", strip=True)


def test_verify_package_requires_fallback_for_failed_formal_scenarios(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    export_review(task_dir)
    build_standard_delivery(task_dir)
    _mark_formal_scenarios(task_dir)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    task["scenario_statuses"]["patenthub_patents"]["status"] = "permission_required"
    task["scenario_statuses"]["patenthub_patents"]["last_message"] = "PatentHub 需要登录。"
    write_json(task_dir / "task.json", task)

    result = verify_package(task_dir)

    assert result["delivery_artifacts_ready"] is True
    assert result["fallback_ready"] is False
    assert result["scenario_coverage_ready"] is False
    assert result["business_ready"] is False
    assert any("PatentHub" in warning for warning in result["warnings"])


def test_verify_package_accepts_reviewed_complete_offline_package(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    _import_complete_life_science_coverage(task_dir)
    generate_draft_evidence_cards(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    row = 2
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=row, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    import_result = import_review(task_dir, workbook)
    _complete_research_integrity(task_dir)
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert import_result["ok"] is True
    assert result["delivery_artifacts_ready"] is True
    assert result["final_review_ready"] is True
    assert result["search_profile_ready"] is True
    assert result["scenario_coverage_ready"] is True
    assert result["fallback_ready"] is True
    assert result["network_ready"] is True
    assert result["research_integrity_ready"] is True
    assert result["business_ready"] is True

    Path(result["standard_delivery"]["report"]).unlink()
    assert verify_package(task_dir)["business_ready"] is False

    build_standard_delivery(task_dir)
    (task_dir / "knowledge" / "literature_graph.json").unlink()
    assert verify_package(task_dir)["business_ready"] is False


def test_verify_package_does_not_require_ad_sources_for_hcg_project(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, HCG_CONFIRMATIONS)
    _write_hcg_material_and_card(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=2, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)
    warning_text = "\n".join(result["warnings"])
    html = (task_dir / "交付目录" / "00_立项调研综合报告.html").read_text(encoding="utf-8")

    assert "wiley_alz" not in warning_text
    assert "Wiley Alzheimer" not in warning_text
    assert "中华神经科" not in warning_text
    assert "Wiley Alzheimer" not in html
    assert "中华神经科" not in html
    assert result["life_science_coverage"]["required"] is True


def test_hcg_requires_life_science_from_confirmed_profile_not_stale_topic(tmp_path: Path):
    state = init_task("旧标题不应决定项目画像", tmp_path)
    task_dir = Path(state.task_dir)
    update_confirmations(task_dir, HCG_CONFIRMATIONS)
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))

    assert requires_life_science_research(task) is True


def test_generic_ivd_test_kit_conservatively_requires_life_science(tmp_path: Path):
    state = init_task("普通 IVD 检测产品 LSR 安全触发", tmp_path)
    task_dir = Path(state.task_dir)
    update_confirmations(
        task_dir,
        {
            **HCG_CONFIRMATIONS,
            "primary_query": "肌钙蛋白I定量检测试剂盒 化学发光法",
            "english_keywords": "cardiac troponin I quantitative chemiluminescent assay IVD",
            "chinese_synonyms": "",
            "intended_use": "心肌损伤辅助诊断",
            "competitor_scope": "NMPA 已注册同类产品",
        },
    )
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))

    assert requires_life_science_research(task) is True


def test_explicit_registry_only_scope_can_disable_life_science(tmp_path: Path):
    state = init_task("只做注册竞品标准", tmp_path)
    task_dir = Path(state.task_dir)
    update_confirmations(
        task_dir,
        {
            **HCG_CONFIRMATIONS,
            "life_science_required": False,
            "life_science_scope": "只做注册/竞品/标准",
        },
    )
    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))

    assert requires_life_science_research(task) is False


def test_verify_package_requires_life_science_for_biomarker_projects(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=2, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["life_science_coverage"]["required"] is True
    assert result["scenario_coverage_ready"] is False
    assert result["business_ready"] is False
    assert any("尚未导入 life-science-research" in warning for warning in result["warnings"])


def test_verify_package_blocks_shallow_life_science_coverage(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    import_life_science_findings(
        json.loads((task_dir / "task.json").read_text(encoding="utf-8"))["task_id"],
        task_dir,
        [
            {
                "source_database": "UniProt",
                "evidence_lane": "protein",
                "entity": "MAPT",
                "result_summary": "Single protein finding.",
            }
        ],
        query="MAPT Alzheimer disease",
    )
    generate_draft_evidence_cards(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for row in range(2, ws.max_row + 1):
        for header, value in {
            "是否纳入报告": "是",
            "一级标签": "临床意义",
            "证据强度": "moderate",
            "复核状态": "已复核",
        }.items():
            ws.cell(row=row, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["life_science_coverage"]["material_count"] == 1
    assert result["scenario_coverage_ready"] is False
    assert any("插件证据数量不足" in warning for warning in result["warnings"])
    assert any("插件数据库覆盖不足" in warning for warning in result["warnings"])
    assert any("插件证据通道覆盖不足" in warning for warning in result["warnings"])


def test_verify_package_blocks_unresolved_network_after_failed_preflight(tmp_path: Path):
    task_dir = _task_dir(tmp_path)
    update_confirmations(task_dir, FULL_CONFIRMATIONS)
    _write_single_material_and_card(task_dir)
    review = export_review(task_dir)
    _mark_formal_scenarios(task_dir)

    task = json.loads((task_dir / "task.json").read_text(encoding="utf-8"))
    task["scenario_statuses"]["pubmed_literature"]["status"] = "not_started"
    task["scenario_statuses"]["pubmed_literature"]["material_count"] = 0
    task["scenario_statuses"]["pubmed_literature"]["last_message"] = ""
    write_json(task_dir / "task.json", task)
    append_jsonl(
        task_dir / "logs" / "events.jsonl",
        {
            "event": "network_preflight",
            "network_ok": False,
            "probes": [
                {
                    "id": "pubmed",
                    "label_zh": "PubMed",
                    "python_dns_error": "No DNS configuration available",
                }
            ],
        },
    )

    from openpyxl import load_workbook

    workbook = Path(review["review_path"])
    wb = load_workbook(workbook)
    ws = wb["文献"]
    headers = [cell.value for cell in ws[1]]
    for header, value in {
        "是否纳入报告": "是",
        "一级标签": "临床意义",
        "证据强度": "moderate",
        "复核状态": "已复核",
    }.items():
        ws.cell(row=2, column=headers.index(header) + 1, value=value)
    wb.save(workbook)

    assert import_review(task_dir, workbook)["ok"] is True
    build_standard_delivery(task_dir)
    result = verify_package(task_dir)

    assert result["network_ready"] is False
    assert result["business_ready"] is False
    assert result["network_unresolved_scenarios"] == [
        {
            "scenario_id": "pubmed_literature",
            "status": "not_started",
            "material_count": 0,
            "last_message": "",
        }
    ]


def test_update_confirmations_rejects_unknown_keys(tmp_path: Path):
    task_dir = _task_dir(tmp_path)

    with pytest.raises(ValueError, match="Unknown confirmation key: target_market"):
        update_confirmations(task_dir, {"target_market": "中国"})
